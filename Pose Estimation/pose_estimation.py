#사용전에 꼭 cv2, numpy, mediapipe 모듈을 설치하여야 구동이 됩니다.
import cv2
import numpy as np

#시각화 및 엑셀 저장을 위한 모듈
import openpyxl
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D


#mediapipe 설치는 cmd 창에서 다음을 입력
#python -m pip install mediapipe
from mediapipe import solutions
from mediapipe.framework.formats import landmark_pb2

import mediapipe as mp
from mediapipe.tasks import python
from mediapipe.tasks.python import vision

#pose_landmarker_lite.task를 아래 링크에 들어가서 Models 부분에서 다운 받기
#https://developers.google.com/mediapipe/solutions/vision/pose_landmarker
model_path = 'C:\pywork\pose_landmarker_lite.task'

#찾은 Point들을 원본 이미지에 표현
def draw_landmarks_on_image(rgb_image, detection_result):
  pose_landmarks_list = detection_result.pose_landmarks
  annotated_image = np.copy(rgb_image)

  # Loop through the detected poses to visualize.
  for idx in range(len(pose_landmarks_list)):
    pose_landmarks = pose_landmarks_list[idx]

    # Draw the pose landmarks.
    pose_landmarks_proto = landmark_pb2.NormalizedLandmarkList()
    pose_landmarks_proto.landmark.extend([
      landmark_pb2.NormalizedLandmark(x=landmark.x, y=landmark.y, z=landmark.z) for landmark in pose_landmarks
    ])
    solutions.drawing_utils.draw_landmarks(
      annotated_image,
      pose_landmarks_proto,
      solutions.pose.POSE_CONNECTIONS,
      solutions.drawing_styles.get_default_pose_landmarks_style())
  return annotated_image


# STEP 2: Create an PoseLandmarker object.
base_options = python.BaseOptions(model_asset_path=model_path)
options = vision.PoseLandmarkerOptions(
    base_options=base_options,
    output_segmentation_masks=True)
detector = vision.PoseLandmarker.create_from_options(options)

#카메라 사용
cap = cv2.VideoCapture(0)

#사용할 지점 0, 11:16, 23:28
position = ["0 - nose","1 - left eye (inner)","2 - left eye","3 - left eye (outer)","4 - right eye (inner)",
"5 - right eye","6 - right eye (outer)","7 - left ear","8 - right ear","9 - mouth (left)","10 - mouth (right)",
"11 - left shoulder","12 - right shoulder","13 - left elbow","14 - right elbow","15 - left wrist","16 - right wrist",
"17 - left pinky","18 - right pinky","19 - left index","20 - right index","21 - left thumb","22 - right thumb",
"23 - left hip","24 - right hip","25 - left knee","26 - right knee","27 - left ankle","28 - right ankle","29 - left heel",
"30 - right heel","31 - left foot index","32 - right foot index"]

#추출한 Point들을 xlxs에 저장하기 위한 기본 세팅
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=2).value = "x"
sheet.cell(row=1, column=3).value = "y"
sheet.cell(row=1, column=4).value = "z"
x=[]
y=[]
z=[]
for i in range(0,13,1):
    x.append(0)
    y.append(0)
    z.append(0)

#시각화 그래프 세팅
plt.ion()#실시간 업데이트 허용
fig = plt.figure()
ax = fig.add_subplot(111)#1개의 plot인 ax를 (1,1)지점에 구현

while True:
    
    # 카메라에서 프레임 읽기
    ret, frame = cap.read()
    
    # 프레임이 정상적으로 읽어졌는지 확인
    if not ret:
        break

    # STEP 3: Load the input image.
    image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    image_frame = mp.Image(
        image_format=mp.ImageFormat.SRGB,
        data=cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    )
    
    # STEP 4: Detect pose landmarks from the input image.
    detection_result = detector.detect(image_frame)

    # STEP 5: Process the detection result. In this case, visualize it.
    annotated_image = draw_landmarks_on_image(image_frame.numpy_view(), detection_result)
    cv2.imshow("image", cv2.cvtColor(annotated_image, cv2.COLOR_RGB2BGR))

    # STEP 5.1 : result to xlsx
    if(detection_result.pose_landmarks.__len__() != 1):continue #사람이 탐색 되지않으면 아랫부분 시행 중지
    for i in range(0,32,1):
        sheet.cell(row=i + 2, column=1).value = position[i]
        sheet.cell(row=i + 2, column=2).value = detection_result.pose_landmarks[0][i].x
        sheet.cell(row=i + 2, column=3).value = detection_result.pose_landmarks[0][i].y
        sheet.cell(row=i + 2, column=4).value = detection_result.pose_landmarks[0][i].z
        #주요 지점 좌표 저장
        if(i==0):
            x[i] = 2 - detection_result.pose_landmarks[0][i].x
            y[i] = 5-1.5 * detection_result.pose_landmarks[0][i].y
            z[i] = detection_result.pose_landmarks[0][i].z
        elif(i>=11 and i<=16):
            x[i-10] = 2-detection_result.pose_landmarks[0][i].x
            y[i-10] = 5-1.5 * detection_result.pose_landmarks[0][i].y
            z[i-10] = detection_result.pose_landmarks[0][i].z
        elif (i >= 23 and i <= 28):
            x[i - 16] = 2-detection_result.pose_landmarks[0][i].x
            y[i - 16] = 5-1.5 * detection_result.pose_landmarks[0][i].y
            z[i - 16] = detection_result.pose_landmarks[0][i].z
    workbook.save('landmarks.xlsx')

    #5.2 실시간 주요포인트 시연
    ax.clear()#이전에 그린 plot 초기화
    plt.scatter(x,y)#산점도 plot에 그리기
    plt.draw()#현재 plot 시각화
    plt.pause(0.005)#시각화한 plot을 보여주기위해 잠깐대기

    # 'q' 키를 누르면 종료
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# 카메라 객체와 윈도우 창 닫기
cap.release()
cv2.destroyAllWindows()